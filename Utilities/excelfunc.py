import openpyxl
from openpyxl.styles import PatternFill


def rowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    maxrows = sheet.max_row
    return maxrows


def columncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    maxcolumn = sheet.max_column
    return maxcolumn


def extractdata(file, sheetname, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    extract = sheet.cell(row, column).value
    return extract


def insertdata(file, sheetname, row, column, value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row, column).value = value
    workbook.save(file)


def greencolor(file, sheetname, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    green = PatternFill(start_color="60b212", end_color="60b212", fill_type='solid')
    sheet.cell(row, column).fill = green
    workbook.save(file)


def redcolor(file, sheetname, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    red = PatternFill(start_color="ff0000", end_color="ff0000", fill_type='solid')
    sheet.cell(row, column).fill = red
    workbook.save(file)
